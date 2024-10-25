from openpyxl import load_workbook
import random
# Carregar o arquivo Excel
caminho_do_arquivo = 'ContagemPF_Augusto2024.xlsx'
workbook = load_workbook(caminho_do_arquivo)
idDemanda = "4482600"
link_demanda = "https://alm.serpro/ccm/web/projects/Gest%C3%A3o%20de%20Demandas%20Internas#action=com.ibm.team.workitem.viewWorkItem&id=4482600"
tipoDemanda = "Apuração"
solicitanteDemanda = "Marcos Antonio Andrade da Costa"
responsavel = "Luiz Rogerio Bayma Campos"
pf = "4.6"
# Selecionar a planilha (por nome ou índice)
sheet = workbook['DGSL5_2024']  # ou workbook.active para a planilha ativa
link_demanda = f'=HYPERLINK("{link_demanda}", "Youtube")'
# Obter o número total de linhas
total_linhas = sheet.max_row + 1
"""sheet.cell(row=total_linhas, column=1).value = idDemanda
sheet.cell(row=total_linhas, column=2).value = link_demanda
sheet.cell(row=total_linhas, column=3).value = tipoDemanda
sheet.cell(row=total_linhas, column=4).value = solicitanteDemanda
sheet.cell(row=total_linhas, column=5).value = responsavel
sheet.cell(row=total_linhas, column=6).value = pf
sheet.cell(row=total_linhas, column=7).value = 'Concluída'
sheet.cell(row=total_linhas, column=8).value = 'Pronto'"""
print(total_linhas)
workbook.save(caminho_do_arquivo) 
