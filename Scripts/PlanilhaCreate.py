from openpyxl import Workbook
import random

workbook = Workbook()

sheet = workbook.active

sheet.tittle = "Demandas"

headers = ['Nome Produto', 'Valor do Fornecedor', 'Lucratividade (%)', 'Quantidade']

for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num, value=header)

def gerar_nome_produto():
    prefixos = ['Super', 'Mega', 'Ultra', 'Power', 'Eco', 'Max']
    tipos = ['Widget', 'Gadget', 'Device', 'Tool', 'Instrument', 'Appliance']
    sufixos = ['Plus', 'Pro', 'X', '2000', 'Prime', 'Elite']
    return f'{random.choice(prefixos)}, {random.choice(tipos)}, {random.choice(sufixos)}'
  
num_produtos = 50

for row_num in range(2, num_produtos + 2):
    nome_produto = gerar_nome_produto()
    valor_fornecedor = round(random.uniform(10.0 , 500.0), 2) # Round é para arredondar o número 
    lucratividade = random.randint(10, 100)
    quantidade = random.randint(1,100)

    # Escrever os dados na planilha

    sheet.cell(row=row_num, column=1, value=nome_produto)
    sheet.cell(row=row_num, column=2, value=valor_fornecedor)
    sheet.cell(row=row_num, column=3, value=lucratividade)
    sheet.cell(row=row_num, column=4, value=quantidade)


file_path = 'estoque.xlsx'
workbook.save(file_path)
