from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
file_path = 'estoque.xlsx'

workbook = load_workbook(file_path)

sheet = workbook['Sheet']


col_nomeProduto = 1 # A
col_valorFornecedor = 2 # B
col_lucratividade = 3 # C
col_quantidade = 4 # D

col_precoVenda = col_quantidade + 1
col_lucroTotal = col_precoVenda + 1
col_valorTotal = col_lucroTotal + 1

sheet.cell(row=1, column=col_precoVenda, value='Preço de Venda')
sheet.cell(row=1, column=col_lucroTotal, value='Lucro Total')
sheet.cell(row=1, column=col_valorTotal, value='Valor Total')

max_row = sheet.max_row

for row in range(2,max_row + 1):
    cell_valorFornecedor = sheet.cell(row=row, column=col_valorFornecedor)
    cell_Lucratividade = sheet.cell(row=row, column=col_lucratividade)
    cell_quantidade = sheet.cell(row=row, column=col_quantidade)
    cell_precoVenda = sheet.cell(row=row, column=col_precoVenda)

    formula_preco_venda = f'={cell_valorFornecedor.coordinate} * (1+{cell_Lucratividade.coordinate}/100)'
    cell_precoVenda.value = formula_preco_venda

    formula_lucro_total = f'=({cell_precoVenda.coordinate} - {cell_valorFornecedor.coordinate}) * {cell_quantidade.coordinate}'
    sheet.cell(row=row, column=col_lucroTotal).value = formula_lucro_total

    formula_valor_total = f'={cell_precoVenda.coordinate} * {cell_quantidade.coordinate}'
    sheet.cell(row=row, column=col_valorTotal).value = formula_valor_total

linha_total = max_row + 2
sheet.cell(row=linha_total,column=col_nomeProduto, value='Totais Gerais')
sheet.merge_cells(start_row=linha_total, start_column=col_nomeProduto, end_row=linha_total, end_column=col_quantidade)

formula_total_lucro = f'=SUM({get_column_letter(col_lucroTotal)}2:{get_column_letter(col_lucroTotal)}{max_row})'
sheet.cell(row=linha_total, column=col_lucroTotal).value = formula_total_lucro


formula_total_valor = f'=SUM({get_column_letter(col_valorTotal)}2:{get_column_letter(col_valorTotal)}{max_row})'
sheet.cell(row=linha_total, column=col_lucroTotal).value = formula_total_valor


workbook.save(file_path) 